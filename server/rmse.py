import pandas as pd
import statsmodels.api as sm
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from sklearn.metrics import mean_squared_error as mse
from math import sqrt
from statistics import mean
import numpy as np


def number_to_column_letter(n):
    column_letter = ''
    while n > 0:
        n, remainder = divmod(n - 1, 26)
        column_letter = chr(remainder + 65) + column_letter
    return column_letter

def shuffleData(subset_data):
    return subset_data.sample(frac=1).reset_index(drop=True)

def writeTrainData(path, sheet, train_sheet):
    data = pd.read_excel(path, sheet_name=sheet)

    n80 = int(len(data) * 0.8)
    
    train_subset_data = data.iloc[0:n80]

    with pd.ExcelWriter(path, engine="openpyxl", mode="a", if_sheet_exists='replace') as writer:
        shuffleData(train_subset_data).to_excel(writer, sheet_name=train_sheet, index=False)

def writeTestData(path, sheet, test_sheet):
    data = pd.read_excel(path, sheet_name=sheet)
    
    n80 = int(len(data) * 0.8)
    last_valid_row_position = data.last_valid_index()
    n_end = data.index.get_loc(last_valid_row_position)
    print("Last row: ", n_end)
    
    test_subset_data = data.iloc[n80:n_end+1]

    with pd.ExcelWriter(path, engine="openpyxl", mode="a", if_sheet_exists='replace') as writer:
        shuffleData(test_subset_data).to_excel(writer, sheet_name=test_sheet, index=False)


def getCoefficients(path, sheet, y, x):
    file = path
    data = pd.read_excel(file, sheet_name=sheet)

    data_cleaned = data[~data.isna().any(axis=1)]  # Remove rows with any NaN
    data_cleaned = data_cleaned[~np.isinf(data_cleaned).any(axis=1)] 

    independent = data[x]
    dependent = data[y]

    independent = sm.add_constant(independent)
    model = sm.OLS(dependent, independent)
    results = model.fit()

    return results.params.tolist()


def writePredictions(path, sheet, coefficients):
    # Load the Excel file into a pandas DataFrame
    df = pd.read_excel(path, sheet_name=sheet)
    # Add headers for the new columns
    headings = ['Actual', 'Predicted', 'Error', 'Error^2']

    last_col = df.columns[df.notna().any()].max()
    last_col_index = df.columns.get_loc(last_col)

    start_col = last_col_index + 4
    if start_col > len(df.columns):
        start_col = len(df.columns)

    j = 0
    for i in range(start_col, start_col + len(headings)):
        df.insert(i, headings[j], None)
        j += 1

    first_column = df.columns[0]
    row_max = df.shape[0]

    # Actual values
    for i in range(row_max):
        df.iloc[i, start_col] = df.iloc[i, 0] 


    # Predicted values
    for i in range(row_max):
        predicted = coefficients[0]
        for j in range(1, len(coefficients)):
            predicted += coefficients[j] * df.iloc[i, j] 
        df.iloc[i, start_col+1] = predicted

    #error and error square
    for i in range(row_max):
        error = df.iloc[i,start_col+1] - df.iloc[i,start_col]
        
        df.iloc[i, start_col+2] = error
        df.iloc[i, start_col+3] = error*error

    

    with pd.ExcelWriter(path, engine="openpyxl", mode="a", if_sheet_exists='replace') as writer:
        df.to_excel(writer, sheet_name=sheet, index=False)

    print("File Saved")


# MU ERROR SILA ANI IDK WTF WHY
def getMSE(actual, predicted):
    return mse(actual, predicted)


'''def writeRMSE(path, sheet):
    # Read the existing sheet
    df = pd.read_excel(path, sheet_name=sheet)
    
    # Extract actual and predicted values
    actual = df["Actual"].tolist()
    predicted = df["Predicted"].tolist()
    
    # Calculate metrics
    mse = getMSE(actual, predicted)
    rmse = sqrt(mse)
    normalized = (rmse / mean(actual)) * 100

    # Prepare metrics data
    metrics = ['MSE', 'RMSE', 'Normalized RMSE']
    values = [mse, rmse, normalized]

    # Determine the last valid column in the existing sheet
    last_column = df.columns.size  # Number of columns in the existing DataFrame

    # Write metrics to the Excel file
    with pd.ExcelWriter(path, engine='openpyxl', mode='a', if_sheet_exists='overlay') as writer:
        # Write metrics in columns
        for col_offset, (metric, value) in enumerate(zip(metrics, values)):
            writer.sheets[sheet].cell(row=1, column=last_column + col_offset + 1).value = metric

'''

def test10x(path, ref_sheet, train_sheet, test_sheet, dep, indep):
    writeTrainData(path, ref_sheet, train_sheet)
    coeff = getCoefficients(path, train_sheet, dep, indep)
    print(coeff)
    for i in range(1,11): 
        currTest = f"{test_sheet}_{i}"
        writeTestData(path, ref_sheet, currTest)

        writePredictions(path, currTest, coeff)

path = "Model.xlsx"
reference_sheet = "Trimmed_Data"

newTrain = "Train_Data"
newTest = "Test_Data"

y = ["TOTEX"]
x = ["TOINC", "URB","FSIZE", "emp_status"]

test10x(path, reference_sheet, newTrain, newTest, y, x)

'''path = "Model.xlsx"
reference_sheet = "Trimmed_Data"

newTrain = "Train_Data"
newTest = "Test_Data"

writeTrainAndTestData(path, reference_sheet, newTrain, newTest)
    
coeff = getCoefficients(path, newTrain, y, x)
print(coeff)

writePredictions(path, newTest, coeff)'''

#writeRMSE(path, newTest)
    



# Load the Excel file
'''file_path = "Model.xlsx"
new_sheet_name = "test_subset"
y_var = pd.read_excel(file_path, sheet_name="Trimmed_Data", usecols=y)
x_var = pd.read_excel(file_path, sheet_name="Trimmed_Data", usecols=x)

data = pd.read_excel(file_path, sheet_name="Trimmed_Data")

train_subset_data = data.iloc[n_start:n_80]

test_subset_data = data.iloc[n_80 + 1:n_end]


with pd.ExcelWriter(file_path, engine="openpyxl", mode="a") as writer:
    shuffleData(train_subset_data).to_excel(writer, sheet_name=new_sheet_name, index=False)

print(f"Subset of data written to a new sheet: '{new_sheet_name}' in {file_path}")'''